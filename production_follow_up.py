import pandas as pd # type: ignore
import math
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Border, PatternFill, Side # type: ignore
import os
from openpyxl.styles import Font # type: ignore
from openpyxl.styles import Alignment

def setPrecision(value, decimal_places):
    factor = 10 ** decimal_places
    return math.floor(value * factor) / factor
    
data = pd.read_csv('D:/1. Work/1. Daily/Production follow up/data.csv')

today = input("Enter yesterday's date: ")
file_name = "D:/1. Work/1. Daily/Production follow up/01. Jan/" + today + ".xlsx"
file2_name = '//192.168.1.231/Planning Internal/Md. Mahfuzur Rahman/Production follow up/01. Jan/' + str(today) + '.xlsx'

template = pd.read_excel("D:/1. Work/1. Daily/Production follow up/template.xlsx")

units = ['JAL', 'JAL3', 'JFL', 'JKL', 'MFL', 'FFL2', 'JKL-U2', 'LINGERIE']

file_count = sum(1 for file in os.listdir('D:/1. Work/1. Daily/Production follow up/01. Jan/') if os.path.isfile(os.path.join('D:/1. Work/1. Daily/Production follow up/01. Jan/', file)))
# completed_days = int(input("Enter completed days: "))
completed_days = int(file_count) + 1

QC_Pass = []
Accu_QC_Pass = []
target = []
prod_smv = []
prod_hr = []
prod_eff = []

for index, row in data.iterrows():
    unit = row['Prod. Date']
    if unit in units:
        # from data
        QC_Pass.append(row['QC Pass'])
        Accu_QC_Pass.append(int(row['Accu. QC Pass']))
        target.append(int(row['Target']))
        prod_smv.append(setPrecision(row['Accu. SMV'], 2))
        prod_hr.append(row['Accu. W.Hour'])
        prod_eff.append(row['Accu. Efficiency'] * 100)

plan_balance = []
req_prod_upto_yes = []
backlog_upto_yes = []
QC_pass_min = []
plan_achv = []
accu_prod_min = []
total_plan_min = []
remaining_days = []
plan_smv = []
plan_hr = []
plan_eff = []
avg_req_per_plan = []
avg_trg_min = []
req_hr = []
plan_pcs_total = 0 

i = 0
for index, row in template.iterrows():
    unit = row['Production Unit']
    if unit in units:
        plan_balance.append(row['Plan (Pcs)'] - Accu_QC_Pass[i])
        print(unit, row['Plan day'], completed_days)
        req_prod_upto_yes.append(int(row['Plan (Pcs)'] * (completed_days / row['Plan day'])))
        backlog_upto_yes.append(req_prod_upto_yes[-1] - Accu_QC_Pass[i])
        QC_pass_min.append(int(QC_Pass[i] * prod_smv[i]))
        accu_prod_min.append(int(prod_smv[i] * Accu_QC_Pass[i]))
        total_plan_min.append(int(row['Plan SMV'] * req_prod_upto_yes[-1]))
        plan_achv.append(setPrecision((accu_prod_min[-1] / total_plan_min[-1]) * 100, 2))
        remaining_days.append(int(row['Plan day'] - completed_days))
        plan_smv.append(row['Plan SMV'])
        plan_hr.append(row['Plan Hr'])
        plan_eff.append(setPrecision(row['Plan Eff%'] * 100, 2))
        avg_req_per_plan.append(int((plan_balance[i] + QC_Pass[i]) / (remaining_days[i] + 1)))
        avg_trg_min.append(int(avg_req_per_plan[-1] * plan_smv[i]))
        req_hr.append(int(avg_req_per_plan[-1] / plan_hr[i]))
        i += 1
    else:
        plan_pcs_total = row['Plan (Pcs)']

print(plan_smv)
i = 0
for index, row in template.iterrows():
    unit = row['Production Unit']
    if unit in units:
        template.loc[index, 'QC Pass (Pcs)'] = int(QC_Pass[i])
        template.loc[index, 'Accu. QC Pass'] = int(Accu_QC_Pass[i])
        template.loc[index, 'Plan Balance'] = plan_balance[i]
        template.loc[index, 'Required Prod upto yesterday'] = req_prod_upto_yes[i]
        template.loc[index, 'Back Log upto yesterday'] = backlog_upto_yes[i]
        template.loc[index, 'QC Pass in minutes'] = QC_pass_min[i]
        template.loc[index, 'Factory Target'] = int(target[i])
        template.loc[index, 'Plan Achv % (monthly, minute based)'] = str(plan_achv[i]) + '%'
        template.loc[index, 'Produced SMV'] = prod_smv[i]
        template.loc[index, 'Produced Hr'] = prod_hr[i]
        template.loc[index, 'Prod. Eff%'] = str(setPrecision(prod_eff[i], 2)) + '%'
        template.loc[index, 'Accu. Produced Minute'] = accu_prod_min[i]
        template.loc[index, 'Total Plan Minute'] = total_plan_min[i]
        template.loc[index, 'Completed Day'] = int(completed_days)
        template.loc[index, 'Remaining Days'] = remaining_days[i]
        template.loc[index, 'Average Requirement as per Plan'] = avg_req_per_plan[i]
        template.loc[index, 'Average target in minutes'] = avg_trg_min[i]
        template.loc[index, 'Required Hr'] = req_hr[i]
        template.loc[index, 'Plan Eff%'] = str(plan_eff[i]) + '%'
        i += 1
    else:
        template.loc[index, 'QC Pass (Pcs)'] = sum(QC_Pass[:-1])
        template.loc[index, 'Accu. QC Pass'] = sum(Accu_QC_Pass[:-1])
        template.loc[index, 'Plan Balance'] = sum(plan_balance[:-1])
        template.loc[index, 'Required Prod upto yesterday'] = sum(req_prod_upto_yes[:-1])
        template.loc[index, 'Back Log upto yesterday'] = sum(backlog_upto_yes[:-1])
        template.loc[index, 'QC Pass in minutes'] = sum(QC_pass_min[:-1])
        template.loc[index, 'Factory Target'] = sum(target[:-1])
        template.loc[index, 'Plan Achv % (monthly, minute based)'] = str(setPrecision(sum(plan_achv[:-1]) / (len(plan_achv) - 1), 2)) + '%'
        avg_prod_smv = setPrecision(sum(prod_smv[:-1]) / (len(prod_smv) - 1), 2)
        avg_plan_smv = setPrecision(sum(plan_smv[:-1]) / (len(plan_smv) - 1), 2)
        template.loc[index, 'Produced SMV'] = setPrecision(sum(prod_smv[:-1]) / (len(prod_smv) - 1), 2)
        template.loc[index, 'Produced Hr'] = setPrecision(sum(prod_hr[:-1]) / (len(prod_hr) - 1), 2)
        avg_prod_eff = sum(prod_eff[:-1]) / (len(prod_eff) - 1)
        avg_prod_eff = setPrecision(avg_prod_eff, 2)
        template.loc[index, 'Prod. Eff%'] = str(avg_prod_eff) + "%"
        total_acc_prod_min = avg_prod_smv * sum(Accu_QC_Pass[:-1])
        template.loc[index, 'Accu. Produced Minute'] = int(total_acc_prod_min)
        total_total_plan_min = avg_plan_smv * sum(req_prod_upto_yes[:-1])
        template.loc[index, 'Total Plan Minute'] = int(total_total_plan_min)
        template.loc[index, 'Completed Day'] = completed_days
        template.loc[index, 'Remaining Days'] = remaining_days[0]
        template.loc[index, 'Average Requirement as per Plan'] = sum(avg_req_per_plan[:-1])
        template.loc[index, 'Average target in minutes'] = sum(avg_trg_min[:-1])
        template.loc[index, 'Required Hr'] = int(sum(req_hr[:-1]) / (len(req_hr) - 1))
        avg_plan_eff = setPrecision(sum(plan_eff[:-1]) / (len(plan_eff) - 1), 2)
        template.loc[index, 'Plan Eff%'] = str(avg_plan_eff) + '%'

est_val = sum(QC_Pass[:-1])
est_val = int(est_val / 10000)
est_val = est_val * 10000

forecast_col_names = ['Plan Qty', 'Production in Completed days', 'Estimated Production/day', 'Remaining days Possible Production', 'Forecasted production', 'Backlog from plan']

estimated_prod_per_day = []
estimated_prod_per_day.append(est_val)
remaining_days_possible_production = []
remaining_days_possible_production.append(est_val * remaining_days[0])
for i in range(5):
    estimated_prod_per_day.append(estimated_prod_per_day[i] + 35000)
    remaining_days_possible_production.append(estimated_prod_per_day[-1] * remaining_days[0])

forecasted_production = []
for i in remaining_days_possible_production:
    forecasted_production.append(sum(Accu_QC_Pass[:-1]) + i)

backlog_from_plan = []
for i in forecasted_production:
    backlog_from_plan.append(plan_pcs_total - i)

forecast_result = pd.DataFrame()
forecast_result[forecast_col_names[0]] = [plan_pcs_total, '-', '-', '-', '-', '-']
forecast_result[forecast_col_names[1]] = [sum(Accu_QC_Pass[:-1]),  '-', '-', '-', '-', '-']
forecast_result[forecast_col_names[2]] = estimated_prod_per_day
forecast_result[forecast_col_names[3]] = remaining_days_possible_production
forecast_result[forecast_col_names[4]] = forecasted_production
forecast_result[forecast_col_names[5]] = backlog_from_plan

with pd.ExcelWriter(file_name) as writer:
    template.to_excel(writer, sheet_name="Production-Follow-up", index=False)
    forecast_result.to_excel(writer, sheet_name="Forecast", index=False)

# Load workbook
wb = load_workbook(file_name)
ws_sheet1 = wb['Production-Follow-up']
ws_sheet2 = wb['Forecast']


def is_number(cell):
    return isinstance(cell.value, (int, float))

# define Formats
align = Alignment(horizontal='center', vertical='center', wrap_text=True)
comma = '#,##0'
border = Border(left=Side(style='thin', color='FFa4a4a4'),
                right=Side(style='thin', color='FFa4a4a4'),
                bottom=Side(style='thin', color='FFa4a4a4'),
                top=Side(style='thin', color='FFa4a4a4'))
title_border = Border(left=Side(style='thin', color='FFa4a4a4'),
                      right=Side(style='thin', color='FFa4a4a4'),
                      top=Side(style='double'),
                      bottom=Side(style='double'))
fill_color = PatternFill(start_color='FF1b770b', end_color='FFFF00', fill_type='solid')  # Yellow fill

# apply the Formats
count = 0
for row in ws_sheet1.iter_rows():
    if count >= 10:
        break
    for cell in row:
        if count == 0 or count == 8:
            cell.alignment = align
            cell.font = Font(bold=True, color='ffffff')
            cell.fill = fill_color
            cell.border = Border()
            cell.border = title_border
        else:
            cell.border = border

        if is_number(cell) == True and cell.value > 999:
            cell.number_format = comma
    count += 1

count = 0
for row in ws_sheet2.iter_rows():
    if count >= 7:
        break
    for cell in row:
        if count == 0:
            cell.alignment = align
            cell.font = Font(bold=True, color='ffffff')
            cell.fill = fill_color
            cell.border = Border()
            cell.border = title_border
        else:
            cell.border = border

        if is_number(cell) == True and cell.value > 999:
            cell.number_format = comma
    count += 1

ws_sheet1.column_dimensions['A'].width = 11
ws_sheet1.column_dimensions['B'].width = 11
ws_sheet1.column_dimensions['C'].width = 11
ws_sheet1.column_dimensions['D'].width = 11
ws_sheet1.column_dimensions['T'].width = 11
ws_sheet1.column_dimensions['U'].width = 11
ws_sheet1.column_dimensions['F'].width = 11
ws_sheet1.column_dimensions['D'].width = 11
ws_sheet1.column_dimensions['E'].width = 11
ws_sheet1.column_dimensions['I'].width = 11
ws_sheet1.column_dimensions['H'].width = 11
ws_sheet1.column_dimensions['J'].width = 11


ws_sheet2_range = ws_sheet2['A1:F7']
start_row = 14
start_col = 1
for row_idx, row in enumerate(ws_sheet2_range, start=start_row):
    for col_idx, cell in enumerate(row, start=start_col):
        target_cell = ws_sheet1.cell(row=row_idx, column=col_idx, value=cell.value)
        # Copy the cell font
        if cell.has_style:
            target_cell.font = Font(name=cell.font.name, 
                                    size=cell.font.size, 
                                    bold=cell.font.bold, 
                                    italic=cell.font.italic, 
                                    vertAlign=cell.font.vertAlign, 
                                    underline=cell.font.underline, 
                                    strike=cell.font.strike, 
                                    color=cell.font.color)

        # Copy the fill (cell background color)
        if cell.fill is not None:
            target_cell.fill = PatternFill(fill_type=cell.fill.fill_type, 
                                            start_color=cell.fill.start_color, 
                                            end_color=cell.fill.end_color)

        # Copy the border
        if cell.border is not None:
            target_cell.border = Border(left=cell.border.left,
                                            right=cell.border.right,
                                            top=cell.border.top,
                                            bottom=cell.border.bottom)
        # Copy alignment (recreate it)
        if cell.alignment is not None:
            target_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                vertical=cell.alignment.vertical,
                                                wrap_text=cell.alignment.wrap_text,
                                                shrink_to_fit=cell.alignment.shrink_to_fit,
                                                indent=cell.alignment.indent)

        # Copy number format
        if cell.number_format:
            target_cell.number_format = cell.number_format

# add title at the forecast table
title_cell = ws_sheet1['A13']
title_cell.value = "Forecast excluding Lingerie:"
title_cell.font = Font(bold=True, underline='single')
# ws_sheet1.merge_cells('A13:C13')

# add the date at the top
ws_sheet1.insert_rows(1)
ws_sheet1['A1'] = 'Date: '
ws_sheet1['A1'].font = Font(bold=True)
ws_sheet1['B1'] = '=TODAY() - 1'
ws_sheet1['B1'].font = Font(bold=True)
ws_sheet1['B1'].number_format = 'DD/MM/YYYY'

wb.save(file_name)
wb.save(file2_name)