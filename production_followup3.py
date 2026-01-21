from bs4 import BeautifulSoup
import math
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Border, PatternFill, Side # type: ignore
import os
from openpyxl.styles import Font # type: ignore
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
import pandas as pd

def get_confirmation():
    choice = input("Do you want to create yesterday's report? (Press 'N' or 'n' for No, any other key for Yes): ").strip()
    return choice.lower() != 'n'

def setPrecision(value, decimal_places):
    factor = 10 ** decimal_places
    return math.floor(value * factor) / factor
    
# data = pd.read_csv('Report.csv', encoding='ISO-8859-1')
with open("Sewing Diagnostic Summary Report.html", "r", encoding="utf-8") as file:
    html_file = file.read()
html_content = BeautifulSoup(html_file, "html.parser")
rows = []
for row in html_content.find_all('tr'):
    row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
    if len(row_data) != 20:
        continue
    rows.append(row_data)
today = ''
if get_confirmation():
    today = datetime.now() - timedelta(days=1)
    today = today.strftime("%d-%b-%y")  # Example: 10-Mar-25
else:
    date = int(input("Enter how many days ago from today's date you want to get: "))
    date = datetime.now() - timedelta(days=date)
    today = date.strftime("%d-%b-%y")
file_name = "01. Jan/" + today + ".xlsx"
file2_name = '//192.168.1.231/Planning Internal/Md. Mahfuzur Rahman/Production follow up/01. Jan/' + str(today) + '.xlsx'

units = ['SubTotal: JAL', 'SubTotal: JFL', 'SubTotal: JKL', 'SubTotal: MFL', 'SubTotal: FFL2', 'SubTotal: JKL-U2', 'SubTotal: LIN', 'SubTotal: GTAL']

file_count = sum(1 for file in os.listdir('01. Jan/') if os.path.isfile(os.path.join('01. Jan/', file)))

completed_days = int(file_count) + 1
print("completed_days: ", completed_days)

wb = load_workbook("template - Copy.xlsx")
ws = wb["Sheet1"]
i = 0
for row in rows:
    if row[1] in units:
        if row[1] == 'SubTotal: LIN':
            # First Table
            ws['B12'] = int(row[8].replace(',', ''))
            ws['C12'] = int(row[9].replace(',', ''))
            ws['E12'] = float(row[5])
            ws['F12'] = float(row[6])
            ws['G12'] = float(row[10].replace('%', ''))/100
            ws['J12'] = int(row[2].replace(',', ''))
            ws['K12'] = int(row[3].replace(',', ''))
            ws['L12'] = int(row[4].replace(',', ''))
            ws['P12'] = completed_days
            # second Table
            ws['C24'] = int(row[14].replace(',', ''))
            ws['H24'] = float(row[19])
            ws['J24'] = float(row[15])
            ws['L24'] = float(row[16].replace('%', ''))/100

        else:
            # First Table
            ws['B' + str(i + 4)] = int(row[8].replace(',', ''))
            ws['C' + str(i + 4)] = int(row[9].replace(',', ''))
            ws['E' + str(i + 4)] = float(row[5])
            ws['F' + str(i + 4)] = float(row[6])
            ws['G' + str(i + 4)] = float(row[10].replace('%', ''))/100
            ws['J' + str(i + 4)] = int(row[2].replace(',', ''))
            ws['K' + str(i + 4)] = int(row[3].replace(',', ''))
            ws['L' + str(i + 4)] = int(row[4].replace(',', ''))
            ws['P' + str(i + 4)] = completed_days
            # second Table
            ws['C' + str(i + 16)] = int(row[14].replace(',', ''))
            ws['H' + str(i + 16)] = float(row[19])
            ws['J' + str(i + 16)] = float(row[15])
            ws['L' + str(i + 16)] = float(row[16].replace('%', ''))/100
            i += 1

# ws['B2'] = today
wb.save(file_name)
wb.save(file2_name)
wb.close()
if os.path.exists(file_name):
    print("File created successfully.")