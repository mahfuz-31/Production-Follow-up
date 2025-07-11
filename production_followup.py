import pandas as pd # type: ignore
import math
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Border, PatternFill, Side # type: ignore
import os
from openpyxl.styles import Font # type: ignore
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
def get_confirmation():
    choice = input("Do you want to create yesterday's report? (Press 'N' or 'n' for No, any other key for Yes): ").strip()
    return choice.lower() != 'n'

def setPrecision(value, decimal_places):
    factor = 10 ** decimal_places
    return math.floor(value * factor) / factor
    
data = pd.read_csv('data.csv')

today = ''
if get_confirmation():
    today = datetime.now() - timedelta(days=1)
    today = today.strftime("%d-%b-%y")  # Example: 10-Mar-25
else:
    date = int(input("Enter how many days ago from today's date you want to get: "))
    date = datetime.now() - timedelta(days=date)
    today = date.strftime("%d-%b-%y")
file_name = "07. Jul/" + today + ".xlsx"
file2_name = '//192.168.1.231/Planning Internal/Md. Mahfuzur Rahman/Production follow up/07. Jul/' + str(today) + '.xlsx'

units = ['JAL', 'JAL3', 'JFL', 'JKL', 'MFL', 'FFL2', 'JKL-U2', 'GTAL', 'GMT TOTAL:', 'LINGERIE']

file_count = sum(1 for file in os.listdir('07. Jul/') if os.path.isfile(os.path.join('07. Jul/', file)))

completed_days = int(file_count) + 1
print("completed_days: ", completed_days)

wb = load_workbook("template.xlsx")
ws = wb["Sheet1"]
i = 0
for index, row in data.iterrows():
    unit = row['Prod. Date']
    if unit in units:
        ws['B' + str(i + 4)] = row['QC Pass']
        ws['D' + str(i + 4)] = row['Accu. QC Pass']
        ws['K' + str(i + 4)] = row['Target']
        ws['O' + str(i + 4)] = row['Accu. SMV']
        ws['Q' + str(i + 4)] = row['Accu. W.Hour']
        ws['S' + str(i + 4)] = row['Accu. Efficiency']
        ws['W' + str(i + 4)] = completed_days
        i += 1

# date_cell
ws['B2'] = today

wb.save(file_name)
wb.close()
if os.path.exists(file_name):
    print("File created successfully.")